# -*- coding: utf-8 -*-
"""探查"汇率换算"sheet 的 AK/AL 列"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['汇率换算']
print('max_row:', ws.max_row, 'max_col:', ws.max_column)
# 表头 R1-R5 全部
for i in range(1, 6):
    cells = [('%s=%s' % (get_column_letter(j + 1), str(v)[:14])) for j, v in enumerate(
        next(ws.iter_rows(min_row=i, max_row=i, max_col=ws.max_column, values_only=True))) if v is not None]
    print('R%d: %s' % (i, ' | '.join(cells[:30]) if cells else '(空)'))
# AK(37)/AL(38) 数据样本
print('\nAK/AL 数据样本:')
for r in [6, 7, 8, 100, 500, ws.max_row]:
    ak = ws.cell(row=r, column=37).value
    al = ws.cell(row=r, column=38).value
    a = ws.cell(row=r, column=1).value
    print('  R%d: A=%s AK=%s AL=%s' % (r, str(a)[:12] if a is not None else None, ak, al))
wb.close()
