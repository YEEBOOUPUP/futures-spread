# -*- coding: utf-8 -*-
"""探查"日报（国外）"sheet 结构：R1~R40 × 前 40 列"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['日报（国外）']
print('sheet:', ws.title, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, max_col=40, values_only=True)):
    cells = [('%s=%s' % (get_column_letter(j + 1), str(v)[:18])) for j, v in enumerate(row) if v is not None]
    print('R%d: %s' % (i + 1, ' | '.join(cells[:14]) if cells else '(空)'))
wb.close()
