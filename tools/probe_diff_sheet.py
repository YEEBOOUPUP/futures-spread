# -*- coding: utf-8 -*-
"""探查"国际油脂油料相关价差"sheet：表头 + A列日期 + 相关列名"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
COLS = ['Q', 'AW', 'BG', 'BH', 'BL', 'BN', 'BK', 'IS', 'IZ', 'IX', 'GN', 'IY', 'GT', 'CA', 'CB', 'BZ', 'BU']

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['国际油脂油料相关价差']
print('max_row:', ws.max_row, 'max_col:', ws.max_column)
# 表头：前 6 行
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, max_col=80, values_only=True)):
    cells = [('%s=%s' % (get_column_letter(j + 1), str(v)[:14])) for j, v in enumerate(row) if v is not None]
    print('R%d: %s' % (i + 1, ' | '.join(cells[:16]) if cells else '(空)'))
# 涉及列的表头（R1-R3 中找）
print('\n涉及列名:')
for col in COLS:
    idx = openpyxl.utils.column_index_from_string(col)
    for i in range(1, 4):
        v = ws.cell(row=i, column=idx).value
        if v is not None:
            print('  %s: R%d=%r' % (col, i, str(v)[:20]))
            break
    else:
        print('  %s: (表头空)' % col)
# A 列日期样本（4752/5118/5483/5848/6213/6579/6944 行）
print('\nA 列日期样本:')
for r in [1, 2, 4752, 5118, 5483, 5848, 6213, 6579, 6944, 7308, ws.max_row]:
    v = ws.cell(row=r, column=1).value
    print('  R%d: %r' % (r, str(v)[:14] if v is not None else None))
wb.close()
