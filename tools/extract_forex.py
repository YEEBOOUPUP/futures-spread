# -*- coding: utf-8 -*-
"""
extract_forex.py — 从"汇率换算"sheet 提取棕榈油内外套用汇率（AK=日期, AL=汇率 CNY/MYR）
输出：data/forex.json { dates: [...], rate: [...] }
用法：python tools/extract_forex.py
"""
import json
import os
from datetime import datetime

import openpyxl

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
SHEET = '汇率换算'
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, 'data', 'forex.json')


def main():
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    ws = wb[SHEET]
    dates, rates = [], []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=38, values_only=True):
        d = row[36]   # AK 列（37）
        v = row[37]   # AL 列（38）
        if isinstance(d, datetime) and isinstance(v, (int, float)):
            dates.append(d.strftime('%Y-%m-%d'))
            rates.append(float(v))
    wb.close()
    if not dates:
        print('未提取到汇率数据！')
        return
    out = {'dates': dates, 'rate': rates}
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
    print('forex.json: %d 天 %s ~ %s, 最新汇率 %.4f' % (len(dates), dates[0], dates[-1], rates[-1]))


if __name__ == '__main__':
    main()
