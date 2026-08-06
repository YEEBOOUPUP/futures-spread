# -*- coding: utf-8 -*-
"""全量扫描"日报（国外）"非空行，识别标题/表头/数据块边界"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['日报（国外）']
prev_nonempty = 0
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=28, values_only=True)):
    cells = [(get_column_letter(j + 1), v) for j, v in enumerate(row) if v is not None]
    if not cells:
        continue
    # 判断类型：只1个单元格且是文本 → 可能是标题；多个数值+日期 → 表头/数据
    kinds = []
    for col, v in cells:
        if isinstance(v, (int, float)):
            kinds.append('num')
        elif isinstance(v, str):
            kinds.append('str')
        else:
            kinds.append('?')
    gap = i + 1 - prev_nonempty - 1
    prev_nonempty = i + 1
    text = ' | '.join('%s=%s' % (c, str(v)[:12]) for c, v in cells[:10])
    print('R%d (gap%d): %s' % (i + 1, gap, text))
wb.close()
