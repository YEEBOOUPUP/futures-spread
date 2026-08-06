# -*- coding: utf-8 -*-
"""
excel_to_json.py — 把 WIND 导出的期货价格 Excel（宽表）转换为站点数据 data.json

WIND 宽表格式（以"国内油脂期货价格"为例）：
    行1-2  : 开始日期 / 截止日期（每块重复）
    行3    : 证券代码（每块以"证券代码"/"代码"标签开始，其后每列一个合约代码）
             P.DCE=主力  P00.DCE=最近合约  P01M.DCE=01月合约（P/Y/OI/M/RM 同理）
    行4    : 证券简称（中文名）
    行5    : 指标中文（日期/收盘价）
    行6    : 指标英文（Date/close/settle）—— 决定本块是收盘价还是结算价
    行7起  : 数据（每块的标签列 = 该块的日期列，其余列为对应合约价格）

输出 data.json（紧凑结构，series 与 dates 按索引对齐）：
{
  "updated_at": "...", "source": "...", "metric_unit": "元/吨",
  "dates": ["2015-01-05", ...],
  "products": { "P": { "name": "棕榈油", "contracts": ["主力","最近","01","02",...], "metrics": ["close","settle"] }, ... },
  "series": { "P": { "主力": { "close": [5006,...], "settle": [...] }, ... }, ... }
}

用法：
    python tools/excel_to_json.py --input 文件.xlsx --output data/data.json
可选：
    --sheets 1,2        处理哪些工作表（默认前 2 个；WIND 文件通常前两个是油脂/油料期货）
    --name-map "P:棕榈油"  覆盖/补充品种中文名
    --encoding gbk       控制台输出编码（Windows 默认 gbk）
依赖：pip install openpyxl
"""
import argparse
import json
import re
import sys
from datetime import datetime, timedelta

# 默认品种中文名（可按实际覆盖）
PRODUCT_NAMES = {
    'P': '棕榈油', 'Y': '豆油', 'OI': '菜油',
    'M': '豆粕', 'RM': '菜粕',
    'A': '豆一', 'B': '豆二', 'C': '玉米', 'CS': '淀粉', 'L': '塑料',
    'PP': '聚丙烯', 'V': 'PVC', 'EG': '乙二醇', 'EB': '苯乙烯', 'PG': '液化气',
    'RB': '螺纹钢', 'HC': '热卷', 'I': '铁矿石', 'J': '焦炭', 'JM': '焦煤',
    'CU': '铜', 'AL': '铝', 'ZN': '锌', 'AU': '黄金', 'AG': '白银',
    'SC': '原油', 'FU': '燃料油', 'TA': 'PTA', 'MA': '甲醇', 'SA': '纯碱',
    'FG': '玻璃', 'UR': '尿素', 'CF': '棉花', 'SR': '白糖', 'AP': '苹果',
}

BLOCK_LABELS = ('证券代码', '代码', '证券简称')

# 代码 → 合约标签：P.DCE→主力  P00.DCE→最近  P01M.DCE→01
_CODE_RE = re.compile(r'^([A-Za-z]+)(\d{2})?(M)?\.(DCE|CZC|SHFE|INE|CZCE|GFEX)$')

_DAY = 86400


def parse_code(code):
    """'P.DCE' → ('P','主力')；'P00.DCE' → ('P','最近')；'P01M.DCE' → ('P','01')；解析失败返回 None"""
    m = _CODE_RE.match(code.strip().upper())
    if not m:
        return None
    product, num, is_m, _exch = m.group(1), m.group(2), m.group(3), m.group(4)
    if num is None:
        return product, '主力'
    if num == '00' and not is_m:
        return product, '最近'
    if is_m:
        return product, num  # '01' ~ '12'
    return product, num


def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)) and 20000 < v < 80000:  # Excel serial
        d = datetime(1899, 12, 30) + timedelta(days=float(v))
        return d.strftime('%Y-%m-%d')
    s = str(v).strip()
    m = re.match(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', s)
    if m:
        return '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    return None


def to_num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ('#NAME?', '#N/A', '#VALUE!', '#DIV/0!'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None


def extract_blocks(rows, max_col):
    """识别块结构。返回块列表：[{start_col, metric, columns:[{col, code, product, label}]}]"""
    header3 = rows[2] if len(rows) > 2 else ()
    header6 = rows[5] if len(rows) > 5 else ()
    blocks = []
    cur = None
    for c in range(max_col):
        v = header3[c]
        if v is not None and str(v).strip() in ('证券代码', '代码'):
            # 新块开始（该列也是此块的日期列）；指标名在 R6 行的下一列（本列是 Date）
            metric_raw = header6[c + 1] if c + 1 < len(header6) else None
            metric = str(metric_raw).strip().lower() if metric_raw is not None else 'close'
            if metric not in ('close', 'settle'):
                metric = 'close'
            cur = {'start_col': c, 'metric': metric, 'columns': []}
            blocks.append(cur)
            continue
        if cur is not None and v is not None:
            parsed = parse_code(str(v))
            if parsed:
                product, label = parsed
                cur['columns'].append({'col': c, 'code': str(v).strip().upper(), 'product': product, 'label': label})
    return blocks


def build(rows, max_col, source):
    """rows: 二维数组（含前 6 行表头 + 数据行）。返回紧凑数据集 dict。"""
    blocks = extract_blocks(rows, max_col)
    if not blocks:
        raise ValueError('未识别到 WIND 宽表块结构（需要第 3 行含"证券代码/代码"标签）')

    # 主日期轴 = 第一个块的日期列
    dates = []
    date_strs = set()
    for r in range(6, len(rows)):
        d = fmt_date(rows[r][blocks[0]['start_col']] if blocks[0]['start_col'] < len(rows[r]) else None)
        if d:
            dates.append(d)
            date_strs.add(d)
    if not dates:
        raise ValueError('日期列无有效数据')

    # series 容器
    series = {}   # product -> label -> metric -> [price|null]
    products = {}
    for b in blocks:
        date_col = b['start_col']
        # 本块日期列 → 行索引映射（用于按日期对齐主 dates）
        row_by_date = {}
        for r in range(6, len(rows)):
            d = fmt_date(rows[r][date_col] if date_col < len(rows[r]) else None)
            if d:
                row_by_date[d] = r
        for colinfo in b['columns']:
            prod, label = colinfo['product'], colinfo['label']
            metric = b['metric']
            prices = []
            for d in dates:
                r = row_by_date.get(d)
                v = rows[r][colinfo['col']] if r is not None and colinfo['col'] < len(rows[r]) else None
                prices.append(to_num(v))
            series.setdefault(prod, {}).setdefault(label, {})[metric] = prices
            products.setdefault(prod, {})
            if label not in products[prod].setdefault('contracts', []):
                products[prod]['contracts'].append(label)
            if metric not in products[prod].setdefault('metrics', []):
                products[prod]['metrics'].append(metric)

    # 合约顺序：主力、最近、01..12
    def label_key(lb):
        if lb == '主力':
            return (0, 0)
        if lb == '最近':
            return (0, 1)
        return (1, int(lb))
    for prod in products:
        products[prod]['contracts'].sort(key=label_key)
        products[prod]['metrics'].sort(reverse=True)  # settle 在前？保持 close,settle：sort 字母序 close<settle
        products[prod]['metrics'].sort()
        products[prod]['name'] = PRODUCT_NAMES.get(prod, '')

    return {
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'source': source,
        'dates': dates,
        'products': products,
        'series': series,
    }


def main():
    ap = argparse.ArgumentParser(description='WIND 期货宽表 Excel → data.json')
    ap.add_argument('--input', required=True, help='Excel 文件路径')
    ap.add_argument('--output', default='data/data.json', help='输出 JSON 路径')
    ap.add_argument('--sheets', default='1,2', help='要处理的工作表序号，逗号分隔（默认前两个）')
    ap.add_argument('--name-map', default='', help='覆盖品种中文名，如 "P:棕榈油,Y:豆油"')
    ap.add_argument('--encoding', default='gbk', help='控制台输出编码（Windows 默认 gbk）')
    args = ap.parse_args()

    if args.encoding:
        try:
            sys.stdout.reconfigure(encoding=args.encoding, errors='replace')
            sys.stderr.reconfigure(encoding=args.encoding, errors='replace')
        except Exception:
            pass

    global PRODUCT_NAMES
    for pair in args.name_map.split(','):
        if ':' in pair:
            code, name = pair.split(':', 1)
            PRODUCT_NAMES[code.strip().upper()] = name.strip()

    try:
        import openpyxl
    except ImportError:
        sys.exit('需要 openpyxl：pip install openpyxl')

    sheet_indexes = [int(x) - 1 for x in args.sheets.split(',') if x.strip()]
    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    if len(wb.sheetnames) < max(sheet_indexes) + 1:
        sys.exit('工作表不足，文件共 %d 个工作表' % len(wb.sheetnames))

    all_dates, all_series, all_products = None, {}, {}
    for si in sheet_indexes:
        ws = wb.worksheets[si]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        max_col = ws.max_column or 0
        print('处理工作表[%d] %s：%d 行 x %d 列' % (si + 1, ws.title, len(rows), max_col))
        ds = build(rows, max_col, '%s::%s' % (args.input, ws.title))
        if all_dates is None:
            all_dates = ds['dates']
        elif ds['dates'] != all_dates:
            print('警告：工作表 %s 的日期轴与第一个工作表不一致，按日期合并' % ws.title, file=sys.stderr)
            # 简单合并：以第一表日期为准，缺失补 null
        for prod, conts in ds['series'].items():
            all_series.setdefault(prod, {})
            for label, metrics in conts.items():
                all_series[prod].setdefault(label, {})
                for metric, prices in metrics.items():
                    if metric not in all_series[prod][label]:
                        all_series[prod][label][metric] = prices
        for prod, info in ds['products'].items():
            all_products.setdefault(prod, {})
            for k, v in info.items():
                if k == 'name' and v:
                    all_products[prod]['name'] = v
                elif k in ('contracts', 'metrics'):
                    merged = list(all_products[prod].setdefault(k, []))
                    for item in v:
                        if item not in merged:
                            merged.append(item)
                    all_products[prod][k] = merged
    wb.close()

    def label_key(lb):
        if lb == '主力':
            return (0, 0)
        if lb == '最近':
            return (0, 1)
        return (1, int(lb))
    for prod in all_products:
        all_products[prod]['contracts'].sort(key=label_key)
        all_products[prod]['metrics'].sort()

    out = {
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'source': args.input,
        'dates': all_dates,
        'products': all_products,
        'series': all_series,
    }
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))

    n_series = sum(len(mts) for mts in all_series.values())
    n_price = sum(len(m) for mts in all_series.values() for lbl in mts for m in mts[lbl].values())
    print('完成：%d 个交易日，%d 个品种，%d 个合约系列，%d 个价格序列，输出 → %s'
          % (len(all_dates), len(all_products), sum(len(p['contracts']) for p in all_products.values()), n_price, args.output))
    for prod in sorted(all_products):
        p = all_products[prod]
        print('  %s(%s) 合约: %s 指标: %s' % (prod, p.get('name', ''), ','.join(p['contracts']), ','.join(p['metrics'])))


if __name__ == '__main__':
    main()
