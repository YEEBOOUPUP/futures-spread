# -*- coding: utf-8 -*-
"""
restore_core.py — 从 WIND Excel 恢复核心品种（P/Y/OI/M/RM）的完整数据
（2015 起 + 主力/最近合约），覆盖被新浪重建覆盖的版本。只取 close（用户只需收盘价）。
用法：python tools/restore_core.py
"""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import excel_to_json as etj  # noqa: E402
import openpyxl  # noqa: E402

CORE = ['P', 'Y', 'OI', 'M', 'RM']
EXCEL = r'C:\Users\10172\OneDrive\Desktop\临时数据处理\WIND价格-wyb.xlsx'
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE, 'data')
INDEX_FILE = os.path.join(DATA_DIR, 'index.json')


def label_key(lb):
    if lb == '主力':
        return (0, 0)
    if lb == '最近':
        return (0, 1)
    return (1, int(lb))


def main():
    if not os.path.exists(EXCEL):
        sys.exit('找不到 Excel: %s' % EXCEL)
    print('读取 WIND Excel ...')
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    all_series = {}  # prod -> label -> {date: close}
    for si in [0, 1]:
        ws = wb.worksheets[si]
        rows = list(ws.iter_rows(values_only=True))
        max_col = ws.max_column or 1
        blocks = etj.extract_blocks(rows, max_col)
        for b in blocks:
            date_col = b['start_col']
            for r in range(6, len(rows)):
                d = etj.fmt_date(rows[r][date_col] if date_col < len(rows[r]) else None)
                if not d:
                    continue
                for ci in b['columns']:
                    if b['metric'] != 'close':
                        continue
                    prod, label = ci['product'], ci['label']
                    v = rows[r][ci['col']] if ci['col'] < len(rows[r]) else None
                    num = etj.to_num(v)
                    if num is not None:
                        all_series.setdefault(prod, {}).setdefault(label, {})[d] = num
    wb.close()

    index = json.load(open(INDEX_FILE, encoding='utf-8'))
    for prod in CORE:
        if prod not in all_series:
            print('跳过（Excel 无此品种）:', prod)
            continue
        series = all_series[prod]
        dates = sorted(set().union(*[set(v.keys()) for v in series.values()]))
        out = {'dates': dates, 'series': {m: {'close': [series[m].get(d) for d in dates]} for m in series}}
        # 防御性合并：旧文件（新浪版）如有 Excel 之外的日期则并入
        fpath = os.path.join(DATA_DIR, prod + '.json')
        if os.path.exists(fpath):
            old = json.load(open(fpath, encoding='utf-8'))
            if not set(old['dates']).issubset(set(dates)):
                for om, oser in old['series'].items():
                    for i, d in enumerate(old['dates']):
                        if i < len(oser['close']) and oser['close'][i] is not None:
                            series.setdefault(om, {})[d] = oser['close'][i]
                dates = sorted(set().union(*[set(v.keys()) for v in series.values()]))
                out = {'dates': dates, 'series': {m: {'close': [series[m].get(d) for d in dates]} for m in series}}
        with open(fpath, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
        contracts = sorted(series.keys(), key=label_key)
        index['products'].setdefault(prod, {})['contracts'] = contracts
        index['products'][prod]['metrics'] = ['close']
        index['products'][prod]['file'] = prod + '.json'
        print('[%s] %d 天 %s ~ %s, 合约: %s' % (prod, len(dates), dates[0], dates[-1], contracts))
    with open(INDEX_FILE, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, separators=(',', ':'))
    print('核心品种恢复完成')


if __name__ == '__main__':
    main()
