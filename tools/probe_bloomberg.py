# -*- coding: utf-8 -*-
"""探查"国际油脂油料价格（彭博）"sheet：S~AC 列（菜籽数据）结构"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['国际油脂油料价格（彭博）']
print('max_row:', ws.max_row, 'max_col:', ws.max_column)
# R1-R6 表头 + S(19)~AC(29) 列
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, max_col=ws.max_column, values_only=True)):
    cells = [('%s=%s' % (get_column_letter(j + 1), str(v)[:22])) for j, v in enumerate(row) if v is not None]
    print('R%d: %s' % (i + 1, ' | '.join(cells[:24]) if cells else '(空)'))
# 只显示 S-AC 列（19-29）的表头
print('\nS~AC 列表头:')
for col in ['S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']:
    idx = openpyxl.utils.column_index_from_string(col)
    vals = []
    for r in range(1, 7):
        v = ws.cell(row=r, column=idx).value
        vals.append(str(v)[:16] if v is not None else '-')
    print('  %s: R1=%s R2=%s R3=%s R4=%s R5=%s R6=%s' % (col, vals[0], vals[1], vals[2], vals[3], vals[4], vals[5]))
# 数据起始与日期样本
print('\nA 列与 S~AC 数据样本:')
for r in [7, 8, 100, 500, 1000, ws.max_row]:
    row = [ws.cell(row=r, column=c).value for c in range(1, 30)]
    print('  R%d: A=%s S=%s T=%s U=%s V=%s W=%s' % (r, str(row[0])[:14] if row[0] is not None else None,
          str(row[18])[:10] if len(row) > 18 else None, str(row[19])[:10] if len(row) > 19 else None,
          str(row[20])[:10] if len(row) > 20 else None, str(row[21])[:10] if len(row) > 21 else None,
          str(row[22])[:10] if len(row) > 22 else None))
wb.close()
