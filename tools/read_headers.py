# -*- coding: utf-8 -*-
"""读"国际油脂油料相关价差"R1-R3 完整表头"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['国际油脂油料相关价差']
for i in range(1, 4):
    vals = [ws.cell(row=i, column=j).value for j in range(1, ws.max_column + 1)]
    cells = ['%s=%s' % (get_column_letter(j), str(v)[:10]) for j, v in enumerate(vals, 1) if v is not None]
    print('R%d (%d 个非空):' % (i, len(cells)))
    # 分块打印
    for k in range(0, len(cells), 30):
        print('  ', ' | '.join(cells[k:k + 30]))
wb.close()
