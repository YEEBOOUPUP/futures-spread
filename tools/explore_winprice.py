# -*- coding: utf-8 -*-
"""探查 WIND价格-wyb.xlsx 的完整结构"""
import openpyxl
from datetime import datetime

PATH = r'C:\Users\10172\OneDrive\Desktop\临时数据处理\WIND价格-wyb.xlsx'

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print('工作表:', wb.sheetnames)

for si, name in enumerate(wb.sheetnames[:2], 1):
    ws = wb[name]
    print(f'\n========== 工作表[{si}] {name}  dims={ws.max_row}行 x {ws.max_column}列 ==========')
    rows = list(ws.iter_rows(values_only=True))
    maxr, maxc = len(rows), ws.max_column

    # 表头区 1-6 行
    for r in range(min(6, maxr)):
        cells = [(c + 1, v) for c, v in enumerate(rows[r][:maxc]) if v is not None]
        print(f'R{r+1}: {cells}')

    # 数据从第 7 行开始（索引 6）
    data_start = 6
    print(f'\n数据起始行索引: {data_start}（即第 {data_start+1} 行）')
    for i in (0, 1, 2, -3, -2, -1):
        row = rows[data_start + i] if data_start + i < maxr else None
        if row is None:
            continue
        cells = [(c + 1, v) for c, v in enumerate(row[:maxc]) if v is not None]
        print(f'数据行{i:+d} (第{data_start+i+1}行): {cells[:12]}')

    # 列 35/36（块2 日期/价格）的数据行
    if maxc >= 35:
        print('\n块2（列35起）数据行:')
        for i in (0, 1, 2):
            idx = data_start + i
            print(f'  第{idx+1}行 列35-37: {rows[idx][34:37]}')

    # 统计数据区：第7行起，第2列~最后一列
    from collections import Counter
    cnt = Counter()
    sample_texts = []
    for r in range(data_start, maxr):
        row = rows[r]
        for c in range(1, maxc):
            v = row[c]
            if v is None:
                cnt['empty'] += 1
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                cnt['number'] += 1
            elif isinstance(v, datetime):
                cnt['datetime'] += 1
            elif isinstance(v, str):
                cnt['str'] += 1
                if len(sample_texts) < 8:
                    sample_texts.append(v)
            else:
                cnt[type(v).__name__] += 1
    print(f'\n数据区统计(第7行起, 非日期列): {dict(cnt)}')
    if sample_texts:
        print('  文本样本:', sample_texts)

    # 日期列（第1列）检查
    dcol = []
    for r in range(data_start, maxr):
        v = rows[r][0]
        if v is not None:
            dcol.append(v)
    print(f'日期列: {len(dcol)} 个非空 / 总 {maxr - data_start}')
    print('  首日期:', dcol[0], ' 尾日期:', dcol[-1], ' 类型:', type(dcol[0]).__name__)

wb.close()
