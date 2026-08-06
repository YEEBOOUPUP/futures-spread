# -*- coding: utf-8 -*-
"""探查 油脂油料数据库.xlsx 的"国外油脂期货"sheet 结构"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print('sheets:', wb.sheetnames)
ws = wb['国外油脂期货']
print('维度:', ws.max_row, 'x', ws.max_column)

# 读前 10 行，A~AO 列（1~41）
rows = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=41, values_only=True)):
    rows.append(row)
    print('R%d: %s' % (i + 1, [str(v)[:14] if v is not None else '' for v in row[:41]]))

# 数据起始行（第 11 行起）看几行
for i, row in enumerate(ws.iter_rows(min_row=11, max_row=13, max_col=8, values_only=True)):
    print('数据R%d: %s' % (11 + i, [str(v)[:12] if v is not None else '' for v in row[:8]]))
wb.close()
