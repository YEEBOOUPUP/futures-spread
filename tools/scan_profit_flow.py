# -*- coding: utf-8 -*-
"""探查"国外油脂利润流"sheet 结构"""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['国外油脂利润流']
print('sheet:', ws.title, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
prev = 0
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True)):
    cells = [(get_column_letter(j + 1), v) for j, v in enumerate(row) if v is not None]
    if not cells:
        continue
    gap = i + 1 - prev - 1
    prev = i + 1
    text = ' | '.join('%s=%s' % (c, str(v)[:14]) for c, v in cells[:12])
    print('R%d (gap%d): %s' % (i + 1, gap, text))
wb.close()
