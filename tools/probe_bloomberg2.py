# -*- coding: utf-8 -*-
"""探查彭博 sheet 尾部 + 周末粒度（iter_rows 快读）"""
import openpyxl
from datetime import datetime

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['国际油脂油料价格（彭博）']

rows = list(ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=29, values_only=True))
wb.close()
print('数据行:', len(rows))
# A 列
dates = [r[0] for r in rows if r[0] is not None]
print('首/末日期:', dates[0], '~', dates[-1])
# 周末检查（找周六）
for r in rows[:15]:
    d = r[0]
    if d:
        print('  ', d, d.strftime('%A'))
# S(19) X(24) Y(25) 最后有值
for ci, name in [(18, 'S'), (23, 'X'), (24, 'Y')]:
    last = None
    for r in reversed(rows):
        v = r[ci] if ci < len(r) else None
        if v is not None and not (isinstance(v, str) and v.startswith('#')):
            last = (r[0], v)
            break
    print('%s 最后有值: %s' % (name, last))
# 非 #N/A 数值统计（S 列）
cnt = sum(1 for r in rows if isinstance(r[18], (int, float)))
print('S 列数值数:', cnt)
