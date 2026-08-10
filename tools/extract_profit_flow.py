# -*- coding: utf-8 -*-
"""
extract_profit_flow.py — 从"国际油脂油料相关价差"sheet 提取"国际利润流"季节性图数据
数据源：用户 Excel 的"日报（国外）"图表引用（仅该 sheet 的 23 个指标列）
输出：data/profit_flow.json
  { axis: [366 个 MM-DD（2020 闰年日历）],
    indicators: { 列: {name, unit, years: {2020: [366], 2021: [366], ...}} },
    order: [列按图表锚点顺序] }
用法：python tools/extract_profit_flow.py
"""
import json
import os
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
SHEET = '国际油脂油料相关价差'
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, 'data', 'profit_flow.json')

# 按"日报（国外）"图表锚点顺序的指标列（仅引用本 sheet 的）
ORDER = ['CB', 'CA', 'CP', 'CM', 'BY', 'JK', 'CF', 'CC', 'IL', 'ED',
         'Q', 'AW', 'BG', 'BH', 'BL', 'BN', 'BK', 'IS', 'IZ', 'IX', 'GN', 'IY', 'GT']

# 年份起行（A 列）与结束行 —— 改为自动检测全部年份（不再限定 2020 起）
# 2020 闰年横轴（366 天 MM-DD）
AXIS = [(datetime(2020, 1, 1) + __import__('datetime').timedelta(days=i)).strftime('%m-%d') for i in range(366)]


def main():
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    ws = wb[SHEET]
    col_ids = {c: column_index_from_string(c) for c in ORDER}
    max_col = max(col_ids.values())

    # 表头名称/单位
    names, units = {}, {}
    for c in ORDER:
        names[c] = ws.cell(row=2, column=col_ids[c]).value or c
        u = ws.cell(row=3, column=col_ids[c]).value
        units[c] = str(u) if u is not None else ''

    # 读取全部数据：A 列日期 + 指标列（用 'YYYY-MM-DD' 作 key，避免跨年覆盖）
    data = {c: {} for c in ORDER}  # 列 -> {'YYYY-MM-DD': value}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=max_col, values_only=True):
        d = row[0]
        key = d.strftime('%Y-%m-%d') if isinstance(d, datetime) else None
        for c in ORDER:
            v = row[col_ids[c] - 1]
            if isinstance(v, (int, float)) and key:
                data[c][key] = float(v)
    wb.close()

    # 全部年份（数据实际覆盖的年份）
    all_keys = set().union(*[set(v.keys()) for v in data.values()])
    years = sorted(set(k[:4] for k in all_keys))
    print('数据年份: %s ~ %s（共 %d 年）' % (years[0], years[-1], len(years)))

    indicators = {}
    for c in ORDER:
        yrs = {}
        for yr in years:
            arr = [data[c].get('%s-%s' % (yr, dd)) for dd in AXIS]
            yrs[str(yr)] = arr
        indicators[c] = {'name': names[c], 'unit': units[c], 'years': yrs}

    out = {'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
           'axis': AXIS, 'order': ORDER, 'indicators': indicators}
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
    print('已生成 %s' % OUT)
    for c in ORDER:
        print('  %-3s %s (%s)' % (c, indicators[c]['name'], indicators[c]['unit']))


if __name__ == '__main__':
    main()
