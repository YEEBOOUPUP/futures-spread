# -*- coding: utf-8 -*-
"""检查 Excel 国外油脂期货末行：马棕 B/C 列、美豆油 AF/AH 列"""
import openpyxl

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
ws = wb['国外油脂期货']
rows = list(ws.iter_rows(min_row=ws.max_row - 3, max_row=ws.max_row, max_col=42, values_only=True))
for i, row in enumerate(rows):
    print('R%d: A=%s B=%s C=%s | AE=%s AF=%s AH=%s' % (
        ws.max_row - 3 + i,
        row[0], row[1], row[2],
        row[30] if len(row) > 30 else None,
        row[31] if len(row) > 31 else None,
        row[33] if len(row) > 33 else None,
    ))
wb.close()
