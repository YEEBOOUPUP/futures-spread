# -*- coding: utf-8 -*-
"""
extract_canola.py — 从"国际油脂油料价格（彭博）"sheet 提取菜籽数据（S~AC 列）

分组（R4 彭博代码为准）：
  加菜籽（ICE / RS 系列，加元/吨）S~X：活跃 + 01/03/05/07/11 交割
  欧菜籽（IJ 系列，欧元/吨）Y~AC：活跃 + 02/05/08/11 交割（AA 列 R3 名称标注错误，代码 IJMAY1 属欧菜籽）

输出：
  data/RS_CAN.json（加菜籽）、data/RS_EU.json（欧菜籽）
  并更新 data/index.json（品种条目，可参与同品种月差 / 跨品种价差）

用法：python tools/extract_canola.py
"""
import json
import os
from datetime import datetime

import openpyxl

PATH = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
SHEET = '国际油脂油料价格（彭博）'
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE, 'data')
INDEX_FILE = os.path.join(DATA_DIR, 'index.json')

# 列定义（1-based）：代码 -> (标签, 名称)
CANOLA = {  # 加菜籽
    'code': 'RS_CAN', 'name': '加菜籽', 'unit': '加元/吨',
    'cols': {19: ('活跃', 'ICE菜籽活跃合约'), 20: ('01', 'ICE菜籽01交割'), 21: ('03', 'ICE菜籽03交割'),
             22: ('05', 'ICE菜籽05交割'), 23: ('07', 'ICE菜籽07交割'), 24: ('11', 'ICE菜籽11交割')},
}
EURO = {  # 欧菜籽
    'code': 'RS_EU', 'name': '欧菜籽', 'unit': '欧元/吨',
    'cols': {25: ('活跃', '欧洲菜籽活跃合约'), 26: ('02', '欧洲菜籽02交割'), 27: ('05', '欧洲菜籽05交割'),
             28: ('08', '欧洲菜籽08交割'), 29: ('11', '欧洲菜籽11交割')},
}


def label_key(lb):
    if lb == '活跃':
        return (0, 0)
    return (1, int(lb))


def main():
    wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=29, values_only=True))
    wb.close()

    index = json.load(open(INDEX_FILE, encoding='utf-8'))
    for spec in (CANOLA, EURO):
        series = {}  # 标签 -> {date: value}
        for row in rows:
            d = row[0]
            key = d.strftime('%Y-%m-%d') if isinstance(d, datetime) else None
            if not key:
                continue
            for col, (label, _) in spec['cols'].items():
                v = row[col - 1]
                if isinstance(v, (int, float)):
                    series.setdefault(label, {})[key] = float(v)
        dates = sorted(set().union(*[set(v.keys()) for v in series.values()]))
        out = {'dates': dates, 'series': {m: {'close': [series[m].get(dd) for dd in dates]} for m in series}}
        fpath = os.path.join(DATA_DIR, spec['code'] + '.json')
        with open(fpath, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
        contracts = sorted(series.keys(), key=label_key)
        index['products'][spec['code']] = {
            'name': spec['name'], 'contracts': contracts, 'metrics': ['close'],
            'file': spec['code'] + '.json', 'foreign': True,
        }
        print('[%s] %s: %d 天 %s ~ %s, 合约: %s' % (spec['code'], spec['name'], len(dates), dates[0], dates[-1], contracts))
    index['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    with open(INDEX_FILE, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, separators=(',', ':'))
    print('菜籽数据提取完成')


if __name__ == '__main__':
    main()
