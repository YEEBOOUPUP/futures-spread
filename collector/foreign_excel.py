# -*- coding: utf-8 -*-
"""
foreign_excel.py — 从用户 Excel（油脂油料数据库.xlsx "国外油脂期货" sheet）导入外盘月度日线

块结构（宽表）：
  块1 A~N ：马棕 —— A=日期, B=活跃合约(主力), C~N=01~12月连续合约（2005 起）
  块2 P~AB：FCPO 具体月份合约（数据稀疏，忽略）
  块3 AE~AN：美豆油 —— AE=日期, AF=BO.CBT 主力, AG=BO00.CBT 最近, AH~AN=01M/03M/05M/07M/08M/09M/10M/12M（2000 起）

输出：data/F_FCPO.json、data/F_BO.json（与国内格式一致：dates + series{标签}.close），幂等合并
用法：python collector/foreign_excel.py
"""
import json
import os
import sys
from datetime import datetime

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE, 'data')
INDEX_FILE = os.path.join(DATA_DIR, 'index.json')
EXCEL = r'C:\Users\10172\OneDrive\Desktop\油脂产业整理-王一波\数据库\油脂油料数据库.xlsx'
SHEET = '国外油脂期货'


def to_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if s.startswith('#'):
        return None
    if s[:4].isdigit() and len(s) >= 10:
        return s[:10].replace('/', '-')
    return None


def to_num(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.startswith('#'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def main():
    if not os.path.exists(EXCEL):
        sys.exit('找不到 Excel: %s' % EXCEL)
    print('读取 %s ...' % EXCEL)
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit('sheet 不存在: %s' % SHEET)
    ws = wb[SHEET]
    max_col = ws.max_column or 76

    # 定义块：马棕（A 日期，B 主力，C~N 01~12）与 美豆油（AE 日期，AF 主力，AG 最近，AH~AO 月份）
    blocks = [
        {
            'code': 'F_FCPO', 'name': '马棕油', 'date_col': 1,
            'cols': {2: '主力'} | {i: '%02d' % (i - 2) for i in range(3, 15)},   # C(3)=01 ... N(14)=12
            'm_label': '马棕', 'c_label': 'FCPO',
        },
        {
            'code': 'F_BO', 'name': '美豆油', 'date_col': 31,                    # AE
            'cols': {32: '主力', 33: '最近'} | {i: '%02d' % month for i, month in zip(range(34, 42), [1, 3, 5, 7, 8, 9, 10, 12])},
            'm_label': '美豆油', 'c_label': 'BO',
        },
    ]

    index = json.load(open(INDEX_FILE, encoding='utf-8')) if os.path.exists(INDEX_FILE) else {'products': {}}

    for blk in blocks:
        series = {}     # 标签 -> {date: close}
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=max_col, values_only=True):
            d = to_date(row[blk['date_col'] - 1] if blk['date_col'] - 1 < len(row) else None)
            if not d:
                continue
            for col, label in blk['cols'].items():
                if col - 1 >= len(row):
                    continue
                v = to_num(row[col - 1])
                if v is None:
                    continue
                series.setdefault(label, {})[d] = v
        if not series:
            print('警告: %s 无数据' % blk['code'])
            continue

        dates = sorted(set().union(*[set(v.keys()) for v in series.values()]))
        out = {'dates': dates, 'series': {m: {'close': [series[m].get(d) for d in dates]} for m in series}}

        # 幂等合并已有文件（按日期）
        fpath = os.path.join(DATA_DIR, blk['code'] + '.json')
        if os.path.exists(fpath):
            old = json.load(open(fpath, encoding='utf-8'))
            if not set(old['dates']).issubset(set(dates)):
                for om, oser in old['series'].items():
                    for i, d in enumerate(old['dates']):
                        if oser['close'][i] is not None:
                            series.setdefault(om, {})[d] = oser['close'][i]
                dates = sorted(set().union(*[set(v.keys()) for v in series.values()]))
                out = {'dates': dates, 'series': {m: {'close': [series[m].get(d) for d in dates]} for m in series}}
        with open(fpath, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, separators=(',', ':'))
        index['products'][blk['code']] = {
            'name': blk['name'], 'contracts': sorted(series.keys()),
            'metrics': ['close'], 'file': blk['code'] + '.json', 'foreign': True,
        }
        print('[%s] %s: %d 个合约标签, %d 天, %s ~ %s' % (
            blk['code'], blk['name'], len(series), len(dates), dates[0], dates[-1]))

    wb.close()
    index['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    with open(INDEX_FILE, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, separators=(',', ':'))
    print('外盘 Excel 导入完成')


if __name__ == '__main__':
    main()
